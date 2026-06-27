#!/usr/bin/env python3
"""Builds data/monthly-donations.xlsx — edit this file in Excel/Sheets,
keep the same column headers, and the website's chart updates automatically.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

months = [
    "2026-06", "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12",
    "2027-01", "2027-02", "2027-03", "2027-04", "2027-05", "2027-06",
    "2027-07", "2027-08", "2027-09", "2027-10", "2027-11", "2027-12",
]

# edit these values month by month
data = {
    "Food":     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Money":    [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Medicine": [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Work":     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Shelter":  [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
}

wb = Workbook()
ws = wb.active
ws.title = "Monthly Donations"

headers = ["Month", "Food", "Money", "Medicine", "Work", "Shelter"]
ws.append(headers)

header_fill = PatternFill("solid", fgColor="0E4B43")
header_font = Font(color="FFFFFF", bold=True)
for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center")

for i, m in enumerate(months):
    row = [m] + [data[cat][i] for cat in ["Food", "Money", "Medicine", "Work", "Shelter"]]
    ws.append(row)

# widths
widths = [12, 10, 10, 12, 10, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

wb.save("data/monthly-donations.xlsx")
print("saved monthly-donations.xlsx")
